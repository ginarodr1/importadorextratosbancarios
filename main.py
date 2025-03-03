import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import Workbook
import pandas as pd
import csv
import os
import bcrypt
import re

def open_main_window():
    root = tk.Tk()
    app = ImportadorExtratos(root)
    root.mainloop()

class SenhaLogin:
    def __init__(self, root, callback):
        self.root = root
        self.callback = callback
        self.root.title("Login")
        self.root.geometry("247x100")
        self.root.config(bg='#f4f4f4')
        self.root.iconbitmap(r"C:\Users\regina.santos\Desktop\Automação\Judite\icon.ico")

        self.titulosenha = tk.Label(root, text="Importador de Extratos está protegido.", font=("Roboto", 10), bg='#f4f4f4')
        self.titulosenha.grid(row=2, column=0, columnspan=2, pady=5, padx=10, sticky="w")

        self.senha_label = tk.Label(root, text="Senha:", font=("Roboto", 10), bg='#f4f4f4')
        self.senha_label.grid(row=3, column=0, columnspan=2, pady=5, padx=10, sticky="w")

        self.senha_entry = tk.Entry(root, show="*", font=("Roboto", 10), width=20)
        self.senha_entry.grid(row=3, column=0, padx=60, sticky="w")

        self.login_button = tk.Button(root, text="Entrar", command=self.verificar_senha, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.login_button.grid(row=4, column=0, columnspan=1, padx=27, sticky="e")

        self.senha_entry.bind("<Return>", lambda event: self.verificar_senha())

        self.root.after(0, lambda: self.root.focus_force())
        self.senha_entry.focus_set()

        self.senha_hash = b'$2b$12$ZlVoXX9jSkWwL.jkPDVgdOlaN0vSecKPWIl8WoEfiUJLA..1A24WS' #! senha em hash gerada e armazenada

    def verificar_senha(self):
        print("Verificando senha...")
        senha = self.senha_entry.get()
        if bcrypt.checkpw(senha.encode(), self.senha_hash):
            print("Senha correta!")
            self.root.destroy()
            self.callback()
        else:
            print("Senha incorreta!")
            messagebox.showerror("Erro", "Senha incorreta!")
            self.senha_entry.delete(0, tk.END)




class ImportadorExtratos:
    def __init__(self, root):
        self.root = root
        self.root.title("Importador de Extratos Bancários")
        self.root.geometry("1000x600")
        self.root.config(bg='#D9D9D9')
        self.root.state("zoomed")
        self.root.iconbitmap(r"C:\Users\regina.santos\Desktop\Automação\Judite\icon.ico")
        self.root.protocol("WM_DELETE_WINDOW", self.fechar_janela)

        for i in range(6):
            root.grid_columnconfigure(i, weight=0)
        for i in range(6):
            root.grid_rowconfigure(i, weight=0)

        self.janelas_filhas = [] #? armazenar todas as janelas filhas

        #! -------------------- TÍTULO PRINCIPAL -------------------- #

        self.title_label = tk.Label(root, text="Importador de Extratos", font=("Roboto", 17, "bold"), bg='#D9D9D9')
        self.title_label.grid(row=3, column=0, padx=10, sticky="w")

        #! -------------------- BOTÕES PRINCIPAIS -------------------- #

        self.btn_importar = tk.Button(root, text="Importar", command=self.mostrar_selecao_conta, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_importar.grid(row=4, column=0, padx=10, sticky="w")

        self.btn_limpar = tk.Button(root, text="Limpar Tudo", command=self.confirmar_limpar_dados, font=("Roboto", 10), bg="#FF5733", fg="white", relief="groove", padx=5, pady=3)
        self.btn_limpar.grid(row=4, column=0, padx=80, sticky="w")

        self.btn_classificar1p = tk.Button(root, text="Classificar 1P", command=self.classificar_dados, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_classificar1p.grid(row=2, column=3, padx=90, sticky="e")

        self.btn_limpar1p = tk.Button(root, text="Limpar 1P", command=self.limpar_1p, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_limpar1p.grid(row=2, column=3, padx=5, sticky="e")

        self.btn_exportar = tk.Button(root, text="Exportar", command=self.exportar_dados, font=("Roboto", 10), bg="#4CAF50", fg="white", relief="groove", padx=5, pady=3)
        self.btn_exportar.grid(row=2, column=4, sticky="w")

        #! -------------------- CAMPOS DE SALDO -------------------- #
        #* -------------- LANÇAMENTO EXTRATO BANCÁRIO -------------- #

        self.saldo_inicial_label = tk.Label(root, text="Saldo Inicial Importado>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.saldo_inicial_label.grid(row=2, column=1, padx=1, sticky="w")
        self.saldo_inicial_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.saldo_inicial_entry.grid(row=2, column=1, columnspan=2, padx=160, sticky="w")

        self.saldo_final_label = tk.Label(root, text="Saldo Final Importado>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.saldo_final_label.grid(row=3, column=1, padx=3, sticky="w")
        self.saldo_final_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.saldo_final_entry.grid(row=3, column=1, columnspan=2, padx=160, sticky="w")

        self.saldo_final_calculado_label = tk.Label(root, text="Saldo Final Calculado>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.saldo_final_calculado_label.grid(row=4, column=1, padx=3, sticky="w")
        self.saldo_final_calculado_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.saldo_final_calculado_entry.grid(row=4, column=1, columnspan=2, padx=160, sticky="w")

        self.diferenca_label = tk.Label(root, text="Diferença>", font=("Roboto", 11), bg='#D9D9D9', anchor='e')
        self.diferenca_label.grid(row=5, column=1, padx=82, sticky="w")
        self.diferenca_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=17)
        self.diferenca_entry.grid(row=5, column=1, columnspan=2, padx=160, sticky="w")


        self.empresa_label = tk.Label(root, text="Empresa>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.empresa_label.grid(row=3, column=2, padx=51, sticky="w")
        self.empresa_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.empresa_entry.grid(row=3, column=2, padx=125, sticky="w")

        self.conta_label = tk.Label(root, text="Conta>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.conta_label.grid(row=4, column=2, padx=72, sticky="w")
        self.conta_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.conta_entry.grid(row=4, column=2, padx=125, sticky="w")

        self.centro_custo_label = tk.Label(root, text="C/Custo>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.centro_custo_label.grid(row=5, column=2, padx=54, sticky="w")
        self.centro_custo_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.centro_custo_entry.grid(row=5, column=2, padx=125, sticky="w")


        self.banco_label = tk.Label(root, text="Banco>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.banco_label.grid(row=4, column=2, padx=1, sticky="e")
        self.banco_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=20)
        self.banco_entry.grid(row=4, column=3, padx=1, sticky="w")
        
        self.agencia_conta_label = tk.Label(root, text="Agência/Conta>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.agencia_conta_label.grid(row=5, column=2, padx=1, sticky="e")
        self.agencia_conta_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=20)
        self.agencia_conta_entry.grid(row=5, column=3, padx=1, sticky="w")


        self.saldo_final_contabil_label = tk.Label(root, text="Saldo Final Contábil>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.saldo_final_contabil_label.grid(row=4, column=3, columnspan=4, padx=95, sticky="e")
        self.saldo_final_contabil_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.saldo_final_contabil_entry.grid(row=4, column=3, columnspan=4, padx=10, sticky="e")

        self.diferenca_extrato_bancario_label = tk.Label(root, text="Diferença c/Extrato Bancário>", font=("Roboto", 11), bg='#D9D9D9', anchor='w')
        self.diferenca_extrato_bancario_label.grid(row=5, column=3, columnspan=4, padx=95, sticky="e")
        self.diferenca_extrato_bancario_entry = tk.Entry(root, font=("Roboto", 10), bd=1, relief="solid", width=10)
        self.diferenca_extrato_bancario_entry.grid(row=5, column=3, columnspan=4, padx=10, sticky="e")


        self.linhadeajuda_label = tk.Label(root, text="Coluna 0", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=0)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 1", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=1)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 2", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=2)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 3", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=3)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 4", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=4)
        self.linhadeajuda_label = tk.Label(root, text="Coluna 5", fg='#D9D9D9', font=("Roboto", 10), bg='#D9D9D9')
        self.linhadeajuda_label.grid(row=1, column=5)

        #* -------------------- TREEVIEW PARA EXIBIR OS DADOS IMPORTADOS -------------------- #
        self.tree = ttk.Treeview(root, columns=(
            "DataLEB", "DescriçãoLEB", "NDocLEB", "ValorLEB", "SaldoLEB",
            "LançamentoLC", "DataLC", "DébitoLC", "D-C/CLC", "CréditoLC",
            "C-C/CLC", "CNPJLC", "HistóricoLC", "ValorLC"
        ), show="headings")

        #* -------------------- DEFINIR CABEÇALHOS DAS COLUNAS ------------------- #
        self.tree.heading("DataLEB", text="Data")
        self.tree.heading("DescriçãoLEB", text="Descrição")
        self.tree.heading("NDocLEB", text="N° Doc")
        self.tree.heading("ValorLEB", text="Valor")
        self.tree.heading("SaldoLEB", text="Saldo")
        self.tree.heading("LançamentoLC", text="Lançamento")
        self.tree.heading("DataLC", text="Data")
        self.tree.heading("DébitoLC", text="Débito")
        self.tree.heading("D-C/CLC", text="D-C/C")
        self.tree.heading("CréditoLC", text="Crédito")
        self.tree.heading("C-C/CLC", text="C-C/C")
        self.tree.heading("CNPJLC", text="CNPJ")
        self.tree.heading("HistóricoLC", text="Histórico")
        self.tree.heading("ValorLC", text="Valor")

        #* -------------------- AJUSTAR O TAMANHO DAS COLUNAS -------------------- #
        self.tree.column("DataLEB", width=60)
        self.tree.column("DescriçãoLEB", width=150)
        self.tree.column("NDocLEB", width=60)
        self.tree.column("ValorLEB", width=60)
        self.tree.column("SaldoLEB", width=60)
        self.tree.column("LançamentoLC", width=100)
        self.tree.column("DataLC", width=60)
        self.tree.column("DébitoLC", width=80)
        self.tree.column("D-C/CLC", width=80)
        self.tree.column("CréditoLC", width=80)
        self.tree.column("C-C/CLC", width=80)
        self.tree.column("CNPJLC", width=80)
        self.tree.column("HistóricoLC", width=100)
        self.tree.column("ValorLC", width=80)
        self.tree.grid(row=7, column=0, columnspan=6, pady=5, padx=10, sticky="nsew")

        #* -------------------- TORNAR A TREEVIEW EXPANSÍVEL -------------------- #
        root.grid_rowconfigure(7, weight=1)
        root.grid_columnconfigure(0, weight=1)
        root.grid_columnconfigure(1, weight=1)
        root.grid_columnconfigure(2, weight=1)
        root.grid_columnconfigure(3, weight=1)
        root.grid_columnconfigure(4, weight=1)
        root.grid_columnconfigure(5, weight=1)

    def mostrar_selecao_conta(self): #? cria janela toplevel
        janela_selecao_conta = tk.Toplevel(self.root)
        app = TelaSelecaoConta(janela_selecao_conta, self.abrir_explorador_arquivos)
        self.janelas_filhas.append(janela_selecao_conta) #? adiciona janela filha à lista
        #root = tk.Tk()
        #app = TelaSelecaoConta(root, self.abrir_explorador_arquivos)
        #root.mainloop()

    def abrir_tela_nova_empresa(self):
        if not hasattr(self, 'tela_nova_empresa') or not self.tela_nova_empresa.winfo_exists():
            self.abrir_tela_nova_empresa = tk.Toplevel(self.root)
            app = TelaNovaEmpresa(self.abrir_tela_nova_empresa, self.salvar_nova_empresa)
            self.janelas_filhas.append(self.tela_nova_empresa) #? adiciona janela filha à lista

    def abrir_tela_nova_conta(self):
        if not hasattr(self, 'tela_nova_conta') or not self.tela_nova_conta.winfo_exists():
            self.tela_nova_conta = tk.Toplevel(self.root)
            app = TelaNovaConta(self.tela_nova_conta, self.salvar_nova_conta)
            self.janelas_filhas.append(self.tela_nova_conta) #? adiciona janela filha à lista

    def fechar_janela(self):
        for janela in self.janelas_filhas: #? fecha todas as janelas filhas
            if janela.winfo_exists():
                janela.destroy()
        self.root.destroy() #? fecha a janela principal

    def abrir_explorador_arquivos(self, empresa, conta, arquivo):
        self.empresa_entry.delete(0, tk.END)
        self.empresa_entry.insert(0, empresa.split(" - ")[0])
        self.conta_entry.delete(0, tk.END)
        self.conta_entry.insert(0, conta.split(" - ")[0])
        self.banco_entry.delete(0, tk.END)
        self.banco_entry.insert(0, conta.split(" - ")[1])

        agencia = conta.split(" - ")[3]
        conta_bancaria = conta.split(" - ")[4]
        self.agencia_conta_entry.delete(0, tk.END)
        self.agencia_conta_entry.insert(0, f"{agencia}/{conta_bancaria}")

        self.importar_dados_arquivo(arquivo)

    def importar_dados_arquivo(self, arquivo):
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        sheet = wb.active

        saldo_inicial = sheet["F10"].value
        saldo_final = sheet["F212"].value
        saldo_final_calculado = sheet["F212"].value

        self.saldo_inicial_entry.delete(0, tk.END)
        self.saldo_inicial_entry.insert(0, saldo_inicial)

        self.saldo_final_entry.delete(0, tk.END)
        self.saldo_final_entry.insert(0, saldo_final)

        self.saldo_final_calculado_entry.delete(0, tk.END)
        self.saldo_final_calculado_entry.insert(0, saldo_final_calculado)

        dados_importados = []

        for row in range(11, sheet.max_row + 1):
            coluna_a = sheet.cell(row=row, column=1).value
            if coluna_a and isinstance(coluna_a, str) and coluna_a.strip().lower() == "total":
                break

            coluna_a = sheet.cell(row=row, column=1).value  #? data
            coluna_b = sheet.cell(row=row, column=2).value  #? descrição
            coluna_c = sheet.cell(row=row, column=3).value  #? n° Doc
            coluna_d = sheet.cell(row=row, column=4).value  #? débito
            coluna_e = sheet.cell(row=row, column=5).value  #? crédito
            coluna_f = sheet.cell(row=row, column=6).value  #? saldo

            def tratar_valor(valor):
                if valor is None or valor == "":
                    return 0
                valor = str(valor).replace(".", "").replace(",", ".")
                try:
                    return float(valor)
                except ValueError:
                    return 0

            coluna_d = tratar_valor(coluna_d)
            coluna_e = tratar_valor(coluna_e)

            valor_total = coluna_d + coluna_e
            dados_importados.append([
                coluna_a, coluna_b, coluna_c, valor_total, coluna_f,
                "", "", "", "", "", "", "", "", ""
            ])

        df = pd.DataFrame(dados_importados, columns=[
            "DataLEB", "DescriçãoLEB", "NDocLEB", "ValorLEB", "SaldoLEB",
            "LançamentoLC", "DataLC", "DébitoLC", "D-C/CLC", "CréditoLC",
            "C-C/CLC", "CNPJLC", "HistóricoLC", "ValorLC"
        ])

        for i in self.tree.get_children(): #? limpar a treeview e adicionar os dados
            self.tree.delete(i)

        for _, row in df.iterrows():
            self.tree.insert("", "end", values=list(row))

    def confirmar_limpar_dados(self):
        resposta = messagebox.askyesno("Atenção", "Tem certeza que deseja limpar todos os dados?")
        if resposta:
            self.limpar_dados()

    def limpar_dados(self):
        self.saldo_inicial_entry.delete(0, tk.END)
        self.saldo_final_entry.delete(0, tk.END)
        self.saldo_final_calculado_entry.delete(0, tk.END)
        self.empresa_entry.delete(0, tk.END)
        self.conta_entry.delete(0, tk.END)
        self.diferenca_entry.delete(0, tk.END)
        self.banco_entry.delete(0, tk.END)
        self.agencia_conta_entry.delete(0, tk.END)
        self.saldo_final_contabil_entry.delete(0, tk.END)
        self.diferenca_extrato_bancario_entry.delete(0, tk.END)
        for i in self.tree.get_children():
            self.tree.delete(i)

    def exportar_dados(self):
        wb = Workbook()
        ws = wb.active

        #* -------------------- DEFINE OS CABEÇALHOS DAS COLUNAS -------------------- #
        colunas_exportar = ["LançamentoLC", "DataLC", "DébitoLC", "D-C/CLC", "CréditoLC",
                      "C-C/CLC", "CNPJLC", "HistóricoLC", "ValorLC"]

        #* -------------------- ITERA SOBRE OS ITENS DA TREEVIEW E ADICIONA À PLANILHA -------------------- #
        dados = []
        for item in self.tree.get_children():
            values = self.tree.item(item, 'values')
            dados.append([values[self.tree["columns"].index(col)] for col in colunas_exportar])

        df = pd.DataFrame(dados, columns=colunas_exportar)

        #* ------------------- SALVA O ARQUIVO EXCEL -------------------- #
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
        if file_path:
            df.to_excel(file_path, index=False)
            #wb.save(file_path)
            messagebox.showinfo("Exportar", "Dados exportados com sucesso!")

    def classificar_dados(self):
        messagebox.showinfo("Classificar 1P", "Essa função não está disponível no momento!")

    def limpar_1p(self):
        messagebox.showinfo("Limpar 1P", "Essa função não está disponível no momento!")

    def fechar_janela(self):
        if hasattr(self, 'tela_selecao_conta') and self.tela_selecao_conta.winfo_exists():
            self.tela_selecao_conta.root.destroy()
        self.root.destroy()




class TelaSelecaoConta:
    def __init__(self, root, callback):
        self.root = root
        self.callback = callback
        self.root.title("Importador")
        self.root.geometry("480x210")
        self.root.config(bg='#f4f4f4')
        self.root.iconbitmap(r"C:\Users\regina.santos\Desktop\Automação\Judite\icon.ico")

        #* -------------------- TÍTULO PRINCIPAL -------------------- #
        self.title_label = tk.Label(root, text="Qual conta bancária irá importar?", font=("Roboto", 17, "bold"), bg='#f4f4f4') #? titulo
        self.title_label.grid(row=1, column=0, columnspan=1, padx=10, pady=10, sticky="w")

        #* -------------------- CAMPOS DE INFORMAÇÕES -------------------- #
        self.label_empresa = tk.Label(root, text="Empresa:", font=("Roboto", 10), bg='#f4f4f4') #? texto "empresa"
        self.label_empresa.grid(row=2, column=0, pady=1, padx=10, sticky="w")

        self.empresas = self.carregar_empresas()
        self.combobox_empresa = ttk.Combobox(root, values=self.empresas, font=("Roboto", 10), width=63)
        self.combobox_empresa.grid(row=3, column=0, pady=1, padx=10, sticky="w")
        self.combobox_empresa.bind("<<ComboboxSelected>>", self.atualizar_contas_contabeis)
        self.combobox_empresa.bind("<KeyRelease>", self.atualizar_contas_contabeis)

        self.label_conta_contabil = tk.Label(root, text="Conta Contábil:", font=("Roboto", 10), bg='#f4f4f4')
        self.label_conta_contabil.grid(row=4, column=0, pady=1, padx=10, sticky="w")

        self.combobox_conta_contabil = ttk.Combobox(root, font=("Roboto", 10), width=63)
        self.combobox_conta_contabil.grid(row=5, column=0, pady=1, padx=10, sticky="w")

        #* -------------------- BOTÕES PRINCIPAIS -------------------- #
        self.btn_nova_empresa = tk.Button(root, text="Nova Empresa", command=self.abrir_tela_nova_empresa, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.btn_nova_empresa.grid(row=7, column=0, pady=15, padx=10, sticky="w")

        self.btn_nova_conta = tk.Button(root, text="Nova Conta", command=self.abrir_tela_nova_conta, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.btn_nova_conta.grid(row=7, column=0, pady=15, padx=109, sticky="w")

        self.btn_alterar_conta = tk.Button(root, text="Alterar Conta", command=self.alterar_conta, font=("Roboto", 10), bg="#FFC107", fg="black")
        self.btn_alterar_conta.grid(row=7, column=0, pady=15, padx=190, sticky="w")

        self.btn_ok = tk.Button(root, text="OK", command=self.confirmar, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.btn_ok.grid(row=7, column=0, columnspan=2, pady=15, padx=83, sticky="e")

        self.btn_cancelar = tk.Button(root, text="Cancelar", command=self.cancelar, font=("Roboto", 10), bg="#FF5733", fg="white")
        self.btn_cancelar.grid(row=7, column=0, columnspan=2, pady=15, padx=15, sticky="e")

    def atualizar_contas_contabeis(self, event=None):
        empresa = self.combobox_empresa.get()
        if empresa:
            codigo_empresa = empresa.split(" - ")[0]
            contas = self.carregar_contas_contabeis(codigo_empresa)
            if contas:
                self.combobox_conta_contabil['values'] = contas
            else:
                self.combobox_conta_contabil['values'] = []
                resposta = messagebox.askyesno("Cadastrar Conta", "Nenhuma conta encontrada para esta empresa. Deseja cadastrar uma nova conta?")
                if resposta:
                    self.abrir_tela_nova_conta()
                else:
                    self.root.lift() #? deixa a janela importador como foco dps
        else:
            self.combobox_conta_contabil['values'] = []

    def carregar_empresas(self):
        empresas = []
        if os.path.exists('empresas.csv'):
            with open('empresas.csv', mode='r', newline='') as file:
                reader = csv.reader(file)
                for row in reader:
                    if row:
                        empresas.append(f"{row[0]} - {row[1]}")
        return empresas

    def carregar_contas_contabeis(self, codigo_empresa):
        contas = []
        if os.path.exists('contas.csv'):
            with open('contas.csv', mode='r', newline='') as file:
                reader = csv.reader(file)
                next(reader, None)
                for row in reader:
                    if row and re.search(rf',{codigo_empresa}$', ','.join(row)):
                        contas.append(f"{row[0]} - {row[1]} - {row[2]} - {row[3]} - {row[4]} - {row[5]}")
        return contas
    

    def abrir_tela_nova_empresa(self):
        if not hasattr(self, 'tela_nova_empresa') or not self.tela_nova_empresa.winfo_exists():
            self.tela_nova_empresa = tk.Toplevel(self.root)
            app = TelaNovaEmpresa(self.tela_nova_empresa, self.salvar_nova_empresa)

    def abrir_tela_nova_conta(self):
        if not hasattr(self, 'tela_nova_conta') or not self.tela_nova_conta.winfo_exists():
            self.tela_nova_conta = tk.Toplevel(self.root)
            app = TelaNovaConta(self.tela_nova_conta, self.salvar_nova_conta)

    def salvar_nova_empresa(self, codigo, razao_social):
        with open('empresas.csv', mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([codigo, razao_social])
        messagebox.showinfo("Nova Empresa", "Empresa adicionada com sucesso!")
        self.combobox_empresa['values'] = self.carregar_empresas()

    def salvar_nova_conta(self, codigo_empresa, banco, agencia, conta_bancaria, conta_ativo, conta_passivo):
        with open('contas.csv', mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([codigo_empresa, banco, agencia, conta_bancaria, conta_ativo, conta_passivo])
        messagebox.showinfo("Nova Conta", "Conta adicionada com sucesso!")
        self.combobox_conta_contabil['values'] = self.carregar_contas_contabeis()

    def verificar_empresa(self, event):
        empresa = self.combobox_empresa.get()
        if empresa:
            codigo = empresa.split(" - ")[0]
            if not any(codigo in e for e in self.empresas):
                self.perguntar_cadastrar_empresa(codigo)
            else:
                self.combobox_conta_contabil.focus_set()

    def verificar_empresa_digitada(self, event):
        empresa = self.combobox_empresa.get()
        if empresa:
            codigo = empresa.split(" - ")[0]
            if not any(codigo in e for e in self.empresas):
                self.perguntar_cadastrar_empresa(codigo)

    def perguntar_cadastrar_empresa(self, codigo):
        resposta = messagebox.askyesno("Cadastrar Empresa", f"A empresa com código {codigo} não está cadastrada. Deseja cadastrá-la?")
        if resposta:
            self.abrir_tela_nova_empresa()

    def confirmar(self):
        empresa = self.combobox_empresa.get()
        conta = self.combobox_conta_contabil.get()
        if empresa and conta:
            pdf_resposta = messagebox.askyesno("Formato do Extrato", "O extrato está em PDF?") #! pergunta se o extrato está em PDF

            arquivo = None
            if pdf_resposta:
                arquivo = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])#! não pergunta sobre xlsx, ajuste do filtro para aparece apenas pdf
            else:
                xlsx_resposta = messagebox.askyesno("Formato do Extrato", "O extrato está em XLSX?") #! pergunta se o extrato está em XLSX

                if xlsx_resposta:
                    arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xls;*.xlsx")]) #! ajuste do filtro para aparece apenas xlsx
                else:
                    messagebox.showinfo("Aviso", "Por favor, selecione um arquivo válido.")

            if arquivo:
                self.callback(empresa, conta, arquivo)  #! chame a função para importar dados do arquivo selecionado

            self.root.destroy()
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha ambos os campos.")

    def cancelar(self):
        self.root.destroy()

    def alterar_conta(self):
        messagebox.showinfo("Alterar Conta", "Funcionalidade de Alterar Conta ainda não implementada.")
        self.root.lift()



        
class TelaNovaConta:
    def __init__(self, root, callback):
        self.root = root
        self.callback = callback
        self.root.title("Nova Conta")
        self.root.geometry("480x240")
        self.root.config(bg='#f4f4f4')
        self.root.iconbitmap(r"C:\Users\regina.santos\Desktop\Automação\Judite\icon.ico")

        #* -------------------- TÍTULO PRINCIPAL -------------------- #
        self.title_label = tk.Label(root, text="Informações da conta bancária", font=("Roboto", 17, "bold"), bg='#f4f4f4') #? titulo
        self.title_label.grid(row=1, column=0, columnspan=1, padx=10, pady=10, sticky="w")

        #* -------------------- CAMPOS DE INFORMAÇÕES -------------------- #
        self.label_codigo_empresa = tk.Label(root, text="Código da Empresa:", font=("Roboto", 10), bg='#f4f4f4') #? texto "codigo da empresa"
        self.label_codigo_empresa.grid(row=2, column=0, pady=1, padx=10, sticky="w")
        self.entry_codigo_empresa = tk.Entry(root, font=("Roboto", 10), width=50)
        self.entry_codigo_empresa.grid(row=3, column=0, pady=1, padx=10, sticky="w")

        self.label_bancos = tk.Label(root, text="Banco:", font=("Roboto", 10), bg='#f4f4f4')
        self.label_bancos.grid(row=4, column=0, pady=1, padx=10, sticky="w")
        self.bancos = ["001 - Banco do Brasil", "077 - Banco Inter", "104 - Banco Caixa Eletrônica", 
                       "237 - Banco Bradesco", "274 - Banco Grafeno", "290 - Banco Pagseguro", 
                       "336 - Banco C6 Bank", "341 - Banco Itau", "353 - Banco Santander", 
                       "399 - Banco HSBC", "422 - Banco Safra", "505 - Credit Suisse", 
                       "707 - Banco Daycoval", "748 - Banco Sicredi"]
        self.combobox_banco = ttk.Combobox(root, values=self.bancos, font=("Roboto", 10), width=62)
        self.combobox_banco.grid(row=5, column=0, pady=1, padx=10, sticky="w")

        self.label_agencia = tk.Label(root, text="Agência:", font=("Roboto", 10), bg='#f4f4f4')
        self.label_agencia.grid(row=6, column=0, pady=1, padx=10, sticky="w")
        self.entry_agencia = tk.Entry(root, font=("Roboto", 10), width=13)
        self.entry_agencia.grid(row=7, column=0, pady=1, padx=10, sticky="w")

        self.label_conta_bancaria = tk.Label(root, text="Conta Bancária:", font=("Roboto", 10), bg='#f4f4f4')
        self.label_conta_bancaria.grid(row=6, column=0, pady=1, padx=116, sticky="w")
        self.entry_conta_bancaria = tk.Entry(root, font=("Roboto", 10), width=15)
        self.entry_conta_bancaria.grid(row=7, column=0, pady=1, padx=116, sticky="w")

        self.label_conta_ativo = tk.Label(root, text="N° Conta Ativo:", font=("Roboto", 10), bg='#f4f4f4')
        self.label_conta_ativo.grid(row=6, column=0, pady=1, padx=240, sticky="w")
        self.entry_conta_ativo = tk.Entry(root, font=("Roboto", 10), width=15)
        self.entry_conta_ativo.grid(row=7, column=0, pady=1, padx=240, sticky="w")

        self.label_conta_passivo = tk.Label(root, text="N° Conta Passivo:", font=("Roboto", 10), bg='#f4f4f4')
        self.label_conta_passivo.grid(row=6, column=0, pady=1, padx=120, sticky="e")
        self.entry_conta_passivo = tk.Entry(root, font=("Roboto", 10), width=15)
        self.entry_conta_passivo.grid(row=7, column=0, pady=1, padx=120, sticky="e")

        #* -------------------- BOTÕES PRINCIPAIS -------------------- #
        self.btn_salvar = tk.Button(root, text="Salvar", command=self.salvar, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.btn_salvar.grid(row=8, column=0, sticky="e", padx=192, pady=8)

        self.btn_cancelar = tk.Button(root, text="Cancelar", command=self.cancelar, font=("Roboto", 10), bg="#FF5733", fg="white")
        self.btn_cancelar.grid(row=8, column=0, sticky="e", padx=122, pady=8)

    def salvar(self):
        codigo_empresa = self.entry_codigo_empresa.get()
        bancos = self.combobox_banco.get()
        agencia = self.entry_agencia.get()
        conta_bancaria = self.entry_conta_bancaria.get()
        conta_ativo = self.entry_conta_ativo.get()
        conta_passivo = self.entry_conta_passivo.get()
        if codigo_empresa and bancos and agencia and conta_bancaria and conta_ativo and conta_passivo:
            self.root.destroy()
            self.callback(conta_ativo, bancos, agencia, conta_bancaria, conta_passivo, codigo_empresa)
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha todos os campos.")

    def cancelar(self):
        self.root.destroy()






class TelaNovaEmpresa:
    def __init__(self, root, callback):
        self.root = root
        self.callback = callback
        self.root.title("Nova Empresa")
        self.root.geometry("480x210")
        self.root.config(bg='#f4f4f4')
        self.root.iconbitmap(r"C:\Users\regina.santos\Desktop\Automação\Judite\icon.ico")

        #* -------------------- TÍTULO PRINCIPAL -------------------- #
        self.title_label = tk.Label(root, text="Informe os dados da empresa", font=("Roboto", 17, "bold"), bg='#f4f4f4') #? titulo
        self.title_label.grid(row=1, column=0, columnspan=1, padx=10, pady=10, sticky="w")

        #* -------------------- CAMPOS DE INFORMAÇÕES -------------------- #
        self.label_codigo = tk.Label(root, text="Código:", font=("Roboto", 10), bg='#f4f4f4') #? texto "codigo"
        self.label_codigo.grid(row=2, column=0, pady=1, padx=10, sticky="w")
        self.entry_codigo = tk.Entry(root, font=("Roboto", 10), width=65) #? caixa
        self.entry_codigo.grid(row=3, column=0, pady=1, padx=10, sticky="w")

        self.label_razao_social = tk.Label(root, text="Razão Social:", font=("Roboto", 10), bg='#f4f4f4') #? texto "razao social"
        self.label_razao_social.grid(row=4, column=0, pady=1, padx=10, sticky="w")
        self.entry_razao_social = tk.Entry(root, font=("Roboto", 10), width=65) #?   caixa
        self.entry_razao_social.grid(row=5, column=0, pady=1, padx=10, sticky="w")

        #* -------------------- BOTÕES PRINCIPAIS -------------------- #
        self.btn_salvar = tk.Button(root, text="Salvar", command=self.salvar, font=("Roboto", 10), bg="#4CAF50", fg="white")
        self.btn_salvar.grid(row=7, column=0, columnspan=2, pady=15, padx=77, sticky="e")

        self.btn_cancelar = tk.Button(root, text="Cancelar", command=self.cancelar, font=("Roboto", 10), bg="#FF5733", fg="white")
        self.btn_cancelar.grid(row=7, column=0, columnspan=2, pady=15, padx=10, sticky="e")

    def salvar(self):
        codigo = self.entry_codigo.get()
        razao_social = self.entry_razao_social.get()
        if codigo and razao_social:
            self.root.destroy()
            self.callback(codigo, razao_social)
        else:
            messagebox.showwarning("Aviso", "Por favor, preencha ambos os campos.")

    def cancelar(self):
        self.root.destroy()


if __name__ == "__main__":
    login_root = tk.Tk()
    login_app = SenhaLogin(login_root, open_main_window)
    login_root.mainloop()
